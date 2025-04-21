import openpyxl
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import font_manager
from scipy.optimize import curve_fit
from openpyxl.drawing.image import Image
import io
import argparse
from pathlib import Path
import sys
import platform
import os

def set_japanese_font():
    """Set appropriate Japanese font based on the system"""
    system_os = platform.system()
    
    if system_os == "Darwin":  # macOS
        jp_font_path = "/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc"  # Hiragino Sans
    elif system_os == "Windows":  # Windows
        jp_font_path = "C:/Windows/Fonts/msgothic.ttc"  # MS Gothic
    else:
        raise EnvironmentError("Unsupported operating system for this script")
    
    # Load and set font
    jp_font = font_manager.FontProperties(fname=jp_font_path)
    plt.rcParams['font.family'] = jp_font.get_name()
    return jp_font

def four_pl(x, A, B, C, D):
    """4-parameter logistic regression"""
    return D + (A-D)/(1.0+((x/C)**B))

def five_pl(x, A, B, C, D, E):
    """5-parameter logistic regression"""
    return D + (A-D)/(1.0+((x/C)**B))**E

def load_data_file(file_path):
    """
    Load data from Excel or CSV file
    
    Parameters:
    -----------
    file_path : str or Path
        Path to input file (Excel or CSV)
    
    Returns:
    --------
    pandas.DataFrame : Loaded data
    """
    file_path = Path(file_path)
    if file_path.suffix.lower() == '.csv':
        # CSV file handling
        try:
            df = pd.read_csv(file_path, header=None)
        except Exception as e:
            raise ValueError(f"Error reading CSV file: {str(e)}")
    elif file_path.suffix.lower() in ['.xlsx', '.xls']:
        # Excel file handling
        try:
            df = pd.read_excel(file_path, header=None)
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")
    else:
        raise ValueError("Unsupported file format. Please use CSV or Excel file.")
    
    return df

def detect_data_structure(df):
    """
    Detect data structure and identify number of dilution columns and valid data range
    
    Parameters:
    -----------
    df : pandas.DataFrame
        Input DataFrame
    
    Returns:
    --------
    tuple : (dilution_rates, num_cols, start_col)
        List of dilution rates, number of columns, data start column number
    """
    # Detect dilution rates from first row
    first_row = df.iloc[0]
    
    # Identify columns that can be converted to numeric values
    numeric_cols = []
    start_col = None
    
    for i, val in enumerate(first_row):
        try:
            if isinstance(val, (int, float)):
                if start_col is None:
                    start_col = i
                numeric_cols.append(i)
            elif isinstance(val, str):
                # Handle formulas starting with '=' or numeric strings
                if val.startswith('=') or val.replace('.', '').isdigit():
                    if start_col is None:
                        start_col = i
                    numeric_cols.append(i)
        except ValueError:
            continue
    
    if not numeric_cols:
        raise ValueError("No dilution rate columns found")
    
    # Identify number of consecutive columns
    num_cols = len(numeric_cols)
    
    # Get dilution rates
    dilution_rates = first_row.iloc[numeric_cols].values
    
    return dilution_rates, num_cols, start_col

def evaluate_dilution_rates(dilution_rates):
    """
    Evaluate dilution rates and convert to numeric values
    
    Parameters:
    -----------
    dilution_rates : array-like
        Array of dilution rates (numeric or formula strings)
    
    Returns:
    --------
    list : List of evaluated dilution rates
    """
    evaluated_rates = []
    for i, rate in enumerate(dilution_rates):
        if isinstance(rate, (int, float)):
            evaluated_rates.append(float(rate))
        elif isinstance(rate, str):
            if rate.startswith('='):
                parts = rate.split('*')
                if len(parts) == 2 and parts[1].isdigit():
                    if i == 0:
                        evaluated_rates.append(float(parts[1]))
                    else:
                        evaluated_rates.append(evaluated_rates[-1] * int(parts[1]))
                else:
                    try:
                        evaluated_rates.append(float(eval(rate[1:])))
                    except:
                        print(f"Warning: Unable to evaluate value '{rate}'")
                        return None
            else:
                try:
                    evaluated_rates.append(float(rate))
                except ValueError:
                    print(f"Warning: Unable to convert value '{rate}' to numeric")
                    return None
        else:
            print(f"Warning: Unknown type value '{rate}'")
            return None
    return evaluated_rates

def calculate_fit_metrics(y_true, y_pred, n_params):
    """
    Calculate fitting quality metrics
    
    Parameters:
    -----------
    y_true : array-like
        Actual values
    y_pred : array-like
        Predicted values
    n_params : int
        Number of model parameters
    
    Returns:
    --------
    dict : Various metrics
    """
    n = len(y_true)
    residuals = y_true - y_pred
    rss = np.sum(residuals**2)
    tss = np.sum((y_true - np.mean(y_true))**2)

    r2 = 1 - (rss/tss)
    adj_r2 = 1 - ((1-r2)*(n-1)/(n-n_params-1))
    aic = n * np.log(rss/n) + 2 * n_params
    bic = n * np.log(rss/n) + n_params * np.log(n)
    rmse = np.sqrt(np.mean(residuals**2))
    
    return {
        'R2': r2,
        'Adjusted_R2': adj_r2,
        'AIC': aic,
        'BIC': bic,
        'RMSE': rmse
    }
    
def get_initial_params(y_data, dilution_rates):
    """
    Optimized initial parameter estimation
    
    Parameters:
    -----------
    y_data : array-like
        Measured value data
    dilution_rates : array-like
        Dilution rates
    
    Returns:
    --------
    dict : Initial parameters
    """
    A_init = np.max(y_data) * 1.05  # Maximum response + 5%
    D_init = np.min(y_data) * 0.95  # Minimum response - 5%
    B_init = 1.0                    # Initial slope coefficient
    
    # EC50 (median) estimation
    mid_response = (A_init + D_init) / 2
    closest_idx = np.argmin(np.abs(y_data - mid_response))
    C_init = dilution_rates[closest_idx]
    
    E_init = 1.0  # Initial asymmetry parameter
    
    return {
        'A': A_init,
        'B': B_init,
        'C': C_init,
        'D': D_init,
        'E': E_init
    }

def fit_curve(x_data, y_data, method, init_params, verbose=False):
    """
    Execute and evaluate curve fitting
    
    Parameters:
    -----------
    x_data : array-like
        x-axis data (dilution rates)
    y_data : array-like
        y-axis data (measured values)
    method : str
        Fitting method ('4' or '5')
    init_params : dict
        Initial parameters
    verbose : bool
        Detailed output flag
    
    Returns:
    --------
    tuple : (optimized parameters, covariance matrix, evaluation metrics, fitted curve)
    """
    try:
        if method == '4':
            bounds = ([0, 0.5, 0, 0], [np.inf, 10, np.inf, np.inf])
            p0 = [init_params['A'], init_params['B'], init_params['C'], init_params['D']]
            popt, pcov = curve_fit(four_pl, x_data, y_data, p0=p0, bounds=bounds, maxfev=50000)
            y_fit = four_pl(x_data, *popt)
            n_params = 4
        else:
            bounds = ([0, 0.5, 0, 0, 0.5], [np.inf, 10, np.inf, np.inf, 5])
            p0 = [init_params['A'], init_params['B'], init_params['C'], init_params['D'], init_params['E']]
            popt, pcov = curve_fit(five_pl, x_data, y_data, p0=p0, bounds=bounds, maxfev=50000)
            y_fit = five_pl(x_data, *popt)
            n_params = 5

        metrics = calculate_fit_metrics(y_data, y_fit, n_params)
        
        if verbose:
            print("\nFitting Results:")
            print(f"  R² = {metrics['R2']:.4f}")
            print(f"  Adjusted R² = {metrics['Adjusted_R2']:.4f}")
            print(f"  RMSE = {metrics['RMSE']:.4e}")
            if metrics['R2'] < 0.99:
                print("  Warning: R² is less than 0.99. Please check the fitting quality.")

        return popt, pcov, metrics, y_fit

    except RuntimeError as e:
        raise RuntimeError(f"Fitting failed: {str(e)}")

def detect_data_blocks(df_data, sample_names, replicates=2, verbose=False, log_print=None):
    """
    Parameters:
    -----------
    df_data : pandas.DataFrame
        Input data frame containing the measurements
    sample_names : array-like
        Array of sample names
    replicates : int
        Number of technical replicates
    verbose : bool
        Detailed output flag
    log_print : function, optional
        Function for logging
    
    Returns:
    --------
    list : List of tuples (start_row, end_row) for each block
    """
    blocks = []
    
    # First, identify only rows with actual sample data
    # Skip NaN or empty sample names
    valid_data_rows = []
    for i, name in enumerate(sample_names):
        if not pd.isna(name) and name != '':
            # Check if the data row has all NaN values
            row_data = pd.to_numeric(df_data.iloc[i], errors='coerce')
            if not row_data.isna().all():
                valid_data_rows.append(i)
    
    if verbose and log_print:
        log_print(f"Number of valid data rows: {len(valid_data_rows)}")
        if valid_data_rows:
            log_print(f"First valid row: {valid_data_rows[0]+1}, Last valid row: {valid_data_rows[-1]+1}")
    
    if not valid_data_rows:
        return []  # Return empty list if no valid data rows
        
    # Block detection
    # Using multiple detection criteria:
    # 1. Blank rows or NaN rows
    # 2. Number of replicates (usually 2 or 3)
    # 3. Plate layout (e.g., 8 rows × 12 columns)
    
    # Split blocks by blank rows
    current_block_start = valid_data_rows[0]
    
    for i in range(1, len(valid_data_rows)):
        curr_row = valid_data_rows[i]
        prev_row = valid_data_rows[i-1]
        
        # Start a new block if rows are not consecutive
        if curr_row > prev_row + 1:
            blocks.append((current_block_start, prev_row))
            current_block_start = curr_row
    
    # Add the last block
    if valid_data_rows:
        blocks.append((current_block_start, valid_data_rows[-1]))
    
    # Check minimum block size
    valid_blocks = []
    for start, end in blocks:
        # Check if the block has at least twice the number of replicates in rows
        # (corresponding to at least 2 samples)
        if end - start + 1 >= replicates * 2:
            valid_blocks.append((start, end))
        elif verbose and log_print:
            log_print(f"Warning: Skipping block that is too small: rows {start+1} to {end+1}")
    
    # If no blocks were detected, treat everything as one block
    if not valid_blocks and valid_data_rows:
        valid_blocks = [(valid_data_rows[0], valid_data_rows[-1])]
        if verbose and log_print:
            log_print("No blocks detected, processing all data as one block")
    
    if verbose and log_print:
        log_print(f"\nDetected data blocks:")
        for i, (start, end) in enumerate(valid_blocks):
            log_print(f"Block {i+1}: rows {start+1} to {end+1}")
            # Show sample names in each block
            sample_in_block = [sample_names[j] for j in range(start, end+1) if not pd.isna(sample_names[j])]
            log_print(f"  Samples: {', '.join(str(s) for s in sample_in_block[:min(5, len(sample_in_block))])}" + 
                    (f" and {len(sample_in_block)-5} more samples" if len(sample_in_block) > 5 else ""))
    
    return valid_blocks

def calculate_titer_with_validation(dilution_rates, y_fit, cutoff, extrapolation_limit=2.0, verbose=False, log_print=None):
    """
    Calculate antibody titer with improved handling of cases where the cutoff line
    does not intersect with the fitted curve within the measured range.
    
    Parameters:
    -----------
    dilution_rates : array-like
        Dilution rates (x-axis data)
    y_fit : array-like
        Fitted curve values (y-axis data)
    cutoff : float
        Cutoff value for determining titer
    extrapolation_limit : float
        Maximum factor for extrapolation (default: 2.0)
    verbose : bool
        Detailed output flag
    log_print : function, optional
        Function for logging
        
    Returns:
    --------
    tuple : (titer, status, message)
        - titer: calculated titer value or limit value
        - status: 'valid', 'below_range', 'above_range', or 'invalid'
        - message: explanatory message about the result
    """
    # Debug information
    if verbose and log_print:
        log_print(f"  Titer validation - Cutoff: {cutoff:.4f}")
        log_print(f"  Y values range from {np.min(y_fit):.4f} to {np.max(y_fit):.4f}")
        log_print(f"  X values range from {np.min(dilution_rates):.1f} to {np.max(dilution_rates):.1f}")
        log_print(f"  First few Y values: {y_fit[0:3]}")
        log_print(f"  Last few Y values: {y_fit[-3:]}")
    
    # Check: Get the last dilution rate and its OD value
    max_dilution = np.max(dilution_rates)
    max_dilution_idx = np.argmax(dilution_rates)
    last_od = y_fit[max_dilution_idx]
    
    # If the last OD value is above the cutoff, the sample never drops below cutoff
    if last_od > cutoff:
        if verbose and log_print:
            log_print(f"  HIGH TITER DETECTED: Last OD ({last_od:.4f}) at dilution {max_dilution} is above cutoff ({cutoff:.4f})")
        titer = max_dilution
        status = 'above_range'
        message = f"OD values never drop below cutoff ({cutoff:.4f}). Even at highest dilution ({max_dilution}), OD is {last_od:.4f}. Titer is reported as >{titer}."
        return titer, status, message
    
    # Check: Get the first dilution rate and its OD value
    min_dilution = np.min(dilution_rates)
    min_dilution_idx = np.argmin(dilution_rates)
    first_od = y_fit[min_dilution_idx]
    
    # If the first OD value is below the cutoff, the sample is already below cutoff at lowest dilution
    if first_od < cutoff:
        if verbose and log_print:
            log_print(f"  LOW TITER DETECTED: First OD ({first_od:.4f}) at dilution {min_dilution} is below cutoff ({cutoff:.4f})")
        titer = min_dilution
        status = 'below_range'
        message = f"OD values are already below cutoff ({cutoff:.4f}) at the lowest dilution ({min_dilution}). Titer is reported as <{titer}."
        return titer, status, message
    
    # Normal case: Find where the curve crosses the cutoff
    # We need to be careful with the order of the data points
    # Sort dilution rates and corresponding fitted values
    sorted_indices = np.argsort(dilution_rates)
    sorted_x = dilution_rates[sorted_indices]
    sorted_y = y_fit[sorted_indices]
    
    # Find indices where y is above and below cutoff
    above_cutoff = sorted_y >= cutoff
    below_cutoff = sorted_y < cutoff
    
    # If no transition is found, use the last boundary point
    if not np.any(below_cutoff) or not np.any(above_cutoff):
        if verbose and log_print:
            log_print(f"  Warning: No clear transition found. This should not happen after the previous checks.")
        
        # If no values below cutoff, report as above_range
        if not np.any(below_cutoff):
            titer = max_dilution
            status = 'above_range'
            message = f"No clear transition point found. Titer is reported as >{titer}."
            return titer, status, message
        
        # If no values above cutoff, report as below_range
        else:
            titer = min_dilution
            status = 'below_range'
            message = f"No clear transition point found. Titer is reported as <{titer}."
            return titer, status, message
    
    # Find the transition point (where the curve crosses the cutoff)
    # Find the first index where y drops below cutoff
    transition_idx = None
    for i in range(len(sorted_y) - 1):
        if sorted_y[i] >= cutoff and sorted_y[i+1] < cutoff:
            transition_idx = i
            break
    
    # If a transition is found, interpolate to find the exact titer
    if transition_idx is not None:
        x1 = sorted_x[transition_idx]
        x2 = sorted_x[transition_idx + 1]
        y1 = sorted_y[transition_idx]
        y2 = sorted_y[transition_idx + 1]
        
        # Linear interpolation
        titer = x1 + (x2 - x1) * (cutoff - y1) / (y2 - y1)
        
        if verbose and log_print:
            log_print(f"  Transition found between dilutions {x1:.1f} and {x2:.1f}")
            log_print(f"  Corresponding OD values: {y1:.4f} and {y2:.4f}")
            log_print(f"  Interpolated titer: {titer:.4f}")
        
        status = 'valid'
        message = f"Titer ({titer:.4f}) is within the measured dilution range."
        return titer, status, message
    
    # Fallback: use numpy's interpolation function
    try:
        if verbose and log_print:
            log_print(f"  Using fallback interpolation method")
        
        # Ensure the arrays are sorted correctly for interpolation
        # Note: np.interp expects x to be increasing
        interp_x = sorted_x
        interp_y = sorted_y
        
        # Find indices where OD values bracket the cutoff
        idx_above = np.where(interp_y >= cutoff)[0]
        idx_below = np.where(interp_y < cutoff)[0]
        
        if len(idx_above) > 0 and len(idx_below) > 0:
            # Find the last point above cutoff and first point below cutoff
            last_above = np.max(idx_above)
            first_below = np.min(idx_below)
            
            # If they are adjacent, we can interpolate
            if first_below == last_above + 1:
                x1 = interp_x[last_above]
                x2 = interp_x[first_below]
                y1 = interp_y[last_above]
                y2 = interp_y[first_below]
                
                # Linear interpolation
                titer = x1 + (x2 - x1) * (cutoff - y1) / (y2 - y1)
                
                if verbose and log_print:
                    log_print(f"  Fallback interpolation successful: titer = {titer:.4f}")
                
                status = 'valid'
                message = f"Titer ({titer:.4f}) is within the measured dilution range."
                return titer, status, message
        
        # If we get here, something unusual happened. Try np.interp directly.
        titer = np.interp(cutoff, sorted_y[::-1], sorted_x[::-1])
        
        if verbose and log_print:
            log_print(f"  Direct np.interp fallback: titer = {titer:.4f}")
        
        status = 'valid'
        message = f"Titer ({titer:.4f}) was calculated using direct interpolation."
        return titer, status, message
        
    except Exception as e:
        if verbose and log_print:
            log_print(f"  Error in interpolation: {str(e)}")
        
        # If all else fails, report the closest dilution rate
        idx = np.argmin(np.abs(sorted_y - cutoff))
        titer = sorted_x[idx]
        status = 'valid'
        message = f"Could not interpolate titer precisely. Closest approximate titer is {titer:.4f}."
        return titer, status, message

def extended_regression_model(x, y, cutoff, model_type='4PL', verbose=False, log_print=None):
    """
    Extend the regression model beyond the measured range to estimate titers
    in cases where the cutoff doesn't intersect with the curve.
    
    Parameters:
    -----------
    x : array-like
        x-axis data (dilution rates)
    y : array-like
        y-axis data (measured values)
    cutoff : float
        Cutoff value
    model_type : str
        Regression model type ('4PL' or '5PL')
    verbose : bool
        Detailed output flag
    log_print : function, optional
        Function for logging
        
    Returns:
    --------
    tuple : (titer, status, message)
        - titer: estimated titer value
        - status: 'extrapolated' or 'failed'
        - message: explanatory message
    """
    try:
        # Debug information
        if verbose and log_print:
            log_print(f"  Attempting extrapolation with {model_type} model")
            log_print(f"  Input data range: x from {np.min(x):.1f} to {np.max(x):.1f}, y from {np.min(y):.4f} to {np.max(y):.4f}")
        
        # Sort input data by dilution rate
        sort_idx = np.argsort(x)
        x_sorted = x[sort_idx]
        y_sorted = y[sort_idx]
        
        # Extend the x-range for extrapolation
        min_x = np.min(x_sorted)
        max_x = np.max(x_sorted)
        
        # Create extended range - go 10x lower and 10x higher
        extended_x = np.logspace(
            np.log10(min_x / 10), 
            np.log10(max_x * 10),
            num=1000
        )
        
        # Get initial parameters
        init_params = get_initial_params(y_sorted, x_sorted)
        
        if verbose and log_print:
            log_print(f"  Initial parameters: A={init_params['A']:.4f}, B={init_params['B']:.4f}, C={init_params['C']:.4f}, D={init_params['D']:.4f}" + 
                     (f", E={init_params['E']:.4f}" if model_type == '5PL' else ""))
        
        # Fit the extended model
        if model_type == '4PL':
            bounds = ([0, 0.5, 0, 0], [np.inf, 10, np.inf, np.inf])
            p0 = [init_params['A'], init_params['B'], init_params['C'], init_params['D']]
            popt, _ = curve_fit(four_pl, x_sorted, y_sorted, p0=p0, bounds=bounds, maxfev=50000)
            extended_y = four_pl(extended_x, *popt)
            if verbose and log_print:
                log_print(f"  Fitted parameters: A={popt[0]:.4f}, B={popt[1]:.4f}, C={popt[2]:.4f}, D={popt[3]:.4f}")
        else:  # 5PL
            bounds = ([0, 0.5, 0, 0, 0.5], [np.inf, 10, np.inf, np.inf, 5])
            p0 = [init_params['A'], init_params['B'], init_params['C'], init_params['D'], init_params['E']]
            popt, _ = curve_fit(five_pl, x_sorted, y_sorted, p0=p0, bounds=bounds, maxfev=50000)
            extended_y = five_pl(extended_x, *popt)
            if verbose and log_print:
                log_print(f"  Fitted parameters: A={popt[0]:.4f}, B={popt[1]:.4f}, C={popt[2]:.4f}, D={popt[3]:.4f}, E={popt[4]:.4f}")
        
        # Check the extended range
        min_extended_y = np.min(extended_y)
        max_extended_y = np.max(extended_y)
        
        if verbose and log_print:
            log_print(f"  Extended model range: y from {min_extended_y:.4f} to {max_extended_y:.4f}")
            
        # Check if the cutoff intersects the extended curve
        # First, are all values above cutoff?
        if min_extended_y > cutoff:
            if verbose and log_print:
                log_print(f"  Extended model never drops below cutoff ({cutoff:.4f})")
            titer = max_x * 10  # Report a very high titer (beyond extrapolation)
            status = 'extrapolated'
            message = f"Even with extrapolation to {titer:.1f}, the curve doesn't drop below cutoff ({cutoff:.4f}). Titer is reported as >{max_x}."
            return titer, status, message
            
        # Second, are all values below cutoff?
        if max_extended_y < cutoff:
            if verbose and log_print:
                log_print(f"  Extended model never rises above cutoff ({cutoff:.4f})")
            titer = min_x / 10  # Report a very low titer (beyond extrapolation)
            status = 'extrapolated'
            message = f"Even with extrapolation to {titer:.1f}, the curve doesn't rise above cutoff ({cutoff:.4f}). Titer is reported as <{min_x}."
            return titer, status, message
            
        # Normal case: Find the transition point in the extended curve
        # Sort the extended data to ensure proper order
        sort_idx = np.argsort(extended_x)
        sorted_extended_x = extended_x[sort_idx]
        sorted_extended_y = extended_y[sort_idx]
        
        # Find the transition point (where the curve crosses the cutoff)
        transition_idx = None
        for i in range(len(sorted_extended_y) - 1):
            if sorted_extended_y[i] >= cutoff and sorted_extended_y[i+1] < cutoff:
                transition_idx = i
                break
        
        # If a transition is found, interpolate to find the exact titer
        if transition_idx is not None:
            x1 = sorted_extended_x[transition_idx]
            x2 = sorted_extended_x[transition_idx + 1]
            y1 = sorted_extended_y[transition_idx]
            y2 = sorted_extended_y[transition_idx + 1]
            
            # Linear interpolation
            titer = x1 + (x2 - x1) * (cutoff - y1) / (y2 - y1)
            
            if verbose and log_print:
                log_print(f"  Transition found in extended model at titer = {titer:.4f}")
            
            # Check if the titer is outside the original measurement range
            if titer > max_x:
                message = f"Extrapolated titer ({titer:.4f}) is above the highest measured dilution ({max_x}). Reported as >{max_x}."
                status = 'extrapolated'
                if verbose and log_print:
                    log_print(f"  Reporting as high titer (>{max_x})")
                return max_x, status, message
            elif titer < min_x:
                message = f"Extrapolated titer ({titer:.4f}) is below the lowest measured dilution ({min_x}). Reported as <{min_x}."
                status = 'extrapolated'
                if verbose and log_print:
                    log_print(f"  Reporting as low titer (<{min_x})")
                return min_x, status, message
            else:
                message = f"Extrapolated titer ({titer:.4f}) is within the measurement range. Extrapolation was successful."
                status = 'extrapolated'
                if verbose and log_print:
                    log_print(f"  Extrapolation successful")
                return titer, status, message
        
        # Fallback: Use numpy's interpolation
        if verbose and log_print:
            log_print(f"  No clear transition found in extended model. Using fallback method.")
            
        # Use np.interp with sorted arrays
        titer = np.interp(cutoff, sorted_extended_y[::-1], sorted_extended_x[::-1])
        
        if titer > max_x:
            message = f"Extrapolated titer ({titer:.4f}) is above the highest measured dilution ({max_x}). Reported as >{max_x}."
            status = 'extrapolated'
            return max_x, status, message
        elif titer < min_x:
            message = f"Extrapolated titer ({titer:.4f}) is below the lowest measured dilution ({min_x}). Reported as <{min_x}."
            status = 'extrapolated'
            return min_x, status, message
        else:
            message = f"Extrapolated titer ({titer:.4f}) was calculated using interpolation in the extended model."
            status = 'extrapolated'
            return titer, status, message
                
    except Exception as e:
        if verbose and log_print:
            log_print(f"  Extrapolation failed: {str(e)}")
        return None, 'failed', f"Extrapolation failed: {str(e)}"

def process_data_and_calculate_titer(file_path, output_path, cutoff, method, replicates=2, verbose=False, log_path=None, allow_extrapolation=False):
    """
    Process ELISA data and calculate titer (version supporting CSV and Excel)
    
    Parameters:
    -----------
    file_path : str
        Input file path (CSV or Excel)
    output_path : str
        Output Excel file path
    cutoff : float
        Cutoff value
    method : str
        Fitting method ('4', '5', 'auto')
    replicates : int
        Number of technical replicates
    verbose : bool
        Detailed output flag
    log_path : str, optional
        Log file path
    allow_extrapolation : bool
        Allow extrapolation beyond measured range
    
    Returns:
    --------
    int : Number of processed samples
    """
    # Open log file
    log_file = open(log_path, 'w', encoding='utf-8') if log_path and verbose else None
    
    def log_print(*args, **kwargs):
        """Internal function for log output"""
        if verbose:
            print(*args, **kwargs)
            if log_file:
                output = ' '.join(str(arg) for arg in args)
                if 'end' in kwargs:
                    output += kwargs['end']
                else:
                    output += '\n'
                log_file.write(output)
                log_file.flush()
                
    try:
        if verbose:
            log_print(f"Processing started: {file_path}")
            log_print(f"File format: {Path(file_path).suffix}")
            log_print(f"Method: {method}PL fitting")
            log_print(f"Cutoff value: {cutoff}")
            log_print(f"Number of technical replicates: {replicates}")
            log_print(f"Allow extrapolation: {allow_extrapolation}")

        # Load data using the new function
        df = load_data_file(file_path)
        
        # Prepare output workbook
        output_wb = openpyxl.Workbook()
        results_sheet = output_wb.active
        results_sheet.title = "Results"
        plots_sheet = output_wb.create_sheet("Plots")

        if verbose:
            log_print("\nData loading details:")
            log_print(f"Total rows: {len(df)}")
            log_print("First few rows:")
            log_print(df.head())

        # Detect data structure
        dilution_rates, num_cols, start_col = detect_data_structure(df)
        
        if verbose:
            log_print(f"\nDetected data structure:")
            log_print(f"Number of columns: {num_cols}")
            log_print(f"Start column: {start_col}")
            log_print(f"Dilution rates: {dilution_rates}")

        # Evaluate dilution rates
        evaluated_rates = evaluate_dilution_rates(dilution_rates)
        if evaluated_rates is None:
            raise ValueError("Dilution rate data contains invalid values")
        
        dilution_rates = np.array(evaluated_rates, dtype=float)

        if verbose:
            log_print(f"Found dilution rates: {dilution_rates}")
            log_print(f"Evaluated dilution rates: {evaluated_rates}")

        # Get sample data (starting from row 1, skipping dilution row)
        sample_names = df.iloc[1:, 0].values
        df_data = df.iloc[1:, start_col:start_col+num_cols]
        
        # Remove any empty rows
        valid_rows = ~pd.isna(sample_names) & (sample_names != '')
        sample_names = sample_names[valid_rows]
        df_data = df_data[valid_rows]

        if verbose:
            log_print("\nData structure after cleanup:")
            log_print(f"Number of data rows: {len(sample_names)}")
            if len(sample_names) > 0:
                log_print(f"First sample name: {sample_names[0]}")
                log_print(f"First row data: {df_data.iloc[0].values}")

        # Use enhanced block detection
        blocks = detect_data_blocks(df_data, sample_names, replicates, verbose, log_print)

        if verbose:
            log_print(f"\nDetected data blocks:")
            for i, (start, end) in enumerate(blocks):
                log_print(f"Block {i+1}: rows {start+1} to {end+1}")

        if not blocks:
            raise ValueError("No valid data blocks found")

        # DataFrame for storing results - Add new columns for titer status
        results_df = pd.DataFrame(columns=[
            'Sample', 'Titer', 'Titer_Status', 'Titer_Note', 'R2', 'Adjusted_R2', 
            'RMSE', 'Fitting_Method'
        ])
        
        # Process each block
        for block_idx, (start_row, end_row) in enumerate(blocks):
            if verbose:
                log_print(f"\nStarting processing of block {block_idx+1}:")
                log_print(f"Row range: {start_row+1} to {end_row+1}")

            block_data = df_data.iloc[start_row:end_row+1]

            # Execute data processing according to number of replicates
            for sample_idx in range(0, end_row - start_row + 1, replicates):
                try:
                    # Calculate row position for plot placement
                    row_position = 30 + (block_idx * 30) + ((sample_idx // replicates) * 30)
                    
                    # Get replicate data and calculate average
                    replicate_data = block_data.iloc[sample_idx:sample_idx+replicates]
                    replicate_numeric = replicate_data.apply(pd.to_numeric, errors='coerce')
                    y_data = replicate_numeric.mean().values

                    if verbose:
                        log_print(f"\n  Processing sample {sample_idx//replicates + 1}:")
                        log_print(f"  Data: {y_data}")

                    if np.isnan(y_data).any():
                        log_print(f"Warning: Sample {sample_idx//replicates + 1} contains invalid data")
                        continue

                    sample_name = sample_names[start_row + sample_idx]

                    if verbose:
                        log_print(f"Processing sample: {sample_name}")
                        process_rows = [start_row + sample_idx + i + 1 for i in range(replicates)]
                        log_print(f"Using data: average of rows {', '.join(map(str, process_rows))}")

                    # Get initial fitting parameters
                    init_params = get_initial_params(y_data, dilution_rates)

                    try:
                        final_method = method
                        if method == 'auto':
                            # Auto selection mode: Compare 4PL and 5PL based on AIC
                            metrics_4pl = None
                            metrics_5pl = None
                            
                            try:
                                popt_4pl, _, metrics_4pl, y_fit_4pl = fit_curve(
                                    dilution_rates, y_data, '4', init_params, verbose
                                )
                            except RuntimeError:
                                if verbose:
                                    log_print("4PL fitting failed")
                            
                            try:
                                popt_5pl, _, metrics_5pl, y_fit_5pl = fit_curve(
                                    dilution_rates, y_data, '5', init_params, verbose
                                )
                            except RuntimeError:
                                if verbose:
                                    log_print("5PL fitting failed")
                            
                            # Select optimal model
                            if metrics_4pl and metrics_5pl:
                                if metrics_4pl['AIC'] < metrics_5pl['AIC']:
                                    popt, metrics, y_fit = popt_4pl, metrics_4pl, y_fit_4pl
                                    final_method = '4'
                                else:
                                    popt, metrics, y_fit = popt_5pl, metrics_5pl, y_fit_5pl
                                    final_method = '5'
                            elif metrics_4pl:
                                popt, metrics, y_fit = popt_4pl, metrics_4pl, y_fit_4pl
                                final_method = '4'
                            elif metrics_5pl:
                                popt, metrics, y_fit = popt_5pl, metrics_5pl, y_fit_5pl
                                final_method = '5'
                            else:
                                raise RuntimeError("Both fitting methods failed")
                        else:
                            # Fit with specified model
                            popt, _, metrics, y_fit = fit_curve(
                                dilution_rates, y_data, method, init_params, verbose
                            )

                        # Calculate titer with enhanced validation
                        titer, titer_status, titer_message = calculate_titer_with_validation(
                            dilution_rates, y_fit, cutoff, verbose=verbose, log_print=log_print
                        )
                        
                        # Try extrapolation if titer is out of range and extrapolation is allowed
                        if allow_extrapolation and titer_status in ['below_range', 'above_range']:
                            if verbose:
                                log_print(f"  Attempting extrapolation for {sample_name}...")
                                
                            extrap_titer, extrap_status, extrap_message = extended_regression_model(
                                dilution_rates, y_data, cutoff, model_type=f"{final_method}PL",
                                verbose=verbose, log_print=log_print
                            )
                            
                            if extrap_status == 'extrapolated':
                                titer = extrap_titer
                                titer_status = 'extrapolated'
                                titer_message = extrap_message
                                if verbose:
                                    log_print(f"  Extrapolation successful: {titer_message}")
                            else:
                                if verbose:
                                    log_print(f"  Extrapolation did not improve result: {extrap_message}")

                        # Format titer display value based on status
                        if titer_status == 'below_range':
                            titer_display = f"<{titer:.1f}"
                        elif titer_status == 'above_range':
                            titer_display = f">{titer:.1f}"
                        else:
                            titer_display = f"{titer:.1f}"
                            
                        if verbose:
                            log_print(f"  Titer calculation result: {titer_display}")
                            log_print(f"  Status: {titer_status}")
                            log_print(f"  Note: {titer_message}")

                        # Add results to DataFrame
                        new_row = pd.DataFrame([{
                            'Sample': sample_name,
                            'Titer': titer,
                            'Titer_Status': titer_status,
                            'Titer_Note': titer_message,
                            'R2': metrics['R2'],
                            'Adjusted_R2': metrics['Adjusted_R2'],
                            'RMSE': metrics['RMSE'],
                            'Fitting_Method': f'{final_method}PL'
                        }])
                        results_df = pd.concat([results_df, new_row], ignore_index=True)

                        # Create graph
                        jp_font = set_japanese_font()
                        plt.figure(figsize=(10, 6))

                        # Calculate standard error for error bars
                        y_err = replicate_numeric.std().values / np.sqrt(replicates)

                        # Get extended x-range for plotting the full curve
                        if titer_status in ['below_range', 'above_range', 'extrapolated']:
                            # Extended x-range for better visualization
                            min_x = np.min(dilution_rates)
                            max_x = np.max(dilution_rates)
                            extended_x = np.logspace(
                                np.log10(min_x / 5), 
                                np.log10(max_x * 5),
                                num=1000
                            )
                            
                            # Extended y values for plotting
                            if final_method == '4':
                                extended_y = four_pl(extended_x, *popt)
                            else:
                                extended_y = five_pl(extended_x, *popt)
                                
                            # Plot the extended curve as a dotted line
                            plt.semilogx(extended_x, extended_y, ':', color='blue', alpha=0.5, label='Extended Curve')
                        
                        # Plot with error bars - actual data points
                        plt.errorbar(dilution_rates, y_data, yerr=y_err, fmt='o', label='Measured Values', capsize=5)
                        
                        # Plot the fitted curve through the measured range
                        plt.semilogx(dilution_rates, y_fit, '-', color='blue', label='Fitting Curve')
                        
                        # Plot cutoff line
                        plt.axhline(y=cutoff, color='r', linestyle='--', label='Cutoff')
                        
                        # Plot titer line - with different styles based on status
                        if titer_status == 'valid':
                            plt.axvline(x=titer, color='g', linestyle='--', label=f'Titer: {titer_display}')
                        elif titer_status == 'extrapolated':
                            plt.axvline(x=titer, color='orange', linestyle='--', label=f'Titer (extrapolated): {titer_display}')
                        elif titer_status == 'below_range':
                            plt.axvline(x=titer, color='purple', linestyle=':', label=f'Titer (below range): {titer_display}')
                        elif titer_status == 'above_range':
                            plt.axvline(x=titer, color='purple', linestyle=':', label=f'Titer (above range): {titer_display}')
                            
                        plt.xlabel('Dilution Rate', fontproperties=jp_font)
                        plt.ylabel('Absorbance', fontproperties=jp_font)
                        plt.title(f'{sample_name} ({final_method}PL Fitting)', fontproperties=jp_font)
                        plt.legend(prop=jp_font)
                        plt.grid(True)
                        
                        # Add titer status annotation to the plot
                        status_text = f"Status: {titer_status}\n"
                        if titer_status != 'valid':
                            status_text += f"Note: {titer_message}"
                            
                        plt.annotate(status_text, xy=(0.02, 0.02), xycoords='axes fraction',
                                    bbox=dict(boxstyle="round,pad=0.3", fc="yellow", alpha=0.3),
                                    fontsize=8, fontproperties=jp_font)

                        # Save plot
                        img_buffer = io.BytesIO()
                        plt.savefig(img_buffer, format='png', dpi=300)
                        img_buffer.seek(0)  # Reset buffer position to the beginning

                        # Save as individual PNG file
                        plot_dir = Path(output_path).parent / 'plots'
                        if not os.path.exists(plot_dir):
                            os.makedirs(plot_dir)
                        plot_path = plot_dir / f'{sample_name}_plot.png'
                        plt.savefig(plot_path, dpi=300, bbox_inches='tight')
                        plt.close()

                        # Place plot in Excel
                        img = Image(img_buffer)
                        img.width = 600
                        img.height = 400
                        
                        plots_sheet.cell(row=row_position-1, column=1, value=sample_name)
                        plots_sheet.add_image(img, f'A{row_position}')

                        if verbose:
                            log_print(f"Plot placement: Sample {sample_name} placed in row {row_position}")

                    except Exception as e:
                        log_print(f"Warning: Error during fitting for {sample_name}: {str(e)}")
                    
                except Exception as e:
                    log_print(f"Warning: Error processing block {block_idx+1}, sample {sample_idx//replicates + 1}: {str(e)}")

        # Write results to Excel sheet
        for i, col in enumerate(results_df.columns):
            results_sheet.cell(row=1, column=i+1, value=col)
        
        for i, row in results_df.iterrows():
            for j, value in enumerate(row):
                results_sheet.cell(row=i+2, column=j+1, value=value)

        # Adjust column width
        for column in results_sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            results_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save workbook
        output_wb.save(output_path)
        return len(results_df)

    except Exception as e:
        raise Exception(f"Error occurred during processing: {str(e)}")

    finally:
        if log_file:
            log_file.close()

def parse_args():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description='ELISA Endpoint Titer Analysis Tool - Enhanced Version with Out-of-Range Detection',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Usage Examples:
  Basic usage (Excel):
    %(prog)s -i data.xlsx -c 0.2
  
  Basic usage (CSV):
    %(prog)s -i data.csv -c 0.2
  
  Single data analysis:
    %(prog)s -i single_data.xlsx -c 0.15 -r 1
  
  Specify 4PL fitting:
    %(prog)s -i data.xlsx -c 0.2 -m 4
    
  Allow extrapolation beyond measured range:
    %(prog)s -i data.xlsx -c 0.2 -e
  
  Output detailed analysis information:
    %(prog)s -i data.xlsx -c 0.2 -v

Input File Format:
  - Excel format (.xlsx) or CSV format (.csv)
  - Row 1: Dilution rates
  - Row 2 and beyond: Sample names and measured values
""")
    
    parser.add_argument('--input', '-i', required=True,
                       help='Input file (Excel or CSV format)')
    
    parser.add_argument('--cutoff', '-c', type=float, required=True,
                       help='Cutoff value')
    
    parser.add_argument('--method', '-m', choices=['4', '5', 'auto'],
                       default='auto',
                       help='Fitting method (4: 4PL, 5: 5PL, auto: automatic selection) Default: auto')
    
    parser.add_argument('--replicates', '-r', type=int, choices=[1, 2],
                       default=2,
                       help='Number of technical replicates (1: single, 2: duplicate) Default: 2')
    
    parser.add_argument('--extrapolation', '-e', action='store_true',
                       help='Allow extrapolation for out-of-range titers')
    
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Display detailed output')
    
    return parser.parse_args()

def main():
    """Main function"""
    args = parse_args()
    
    try:
        input_path = Path(args.input)
        output_path = input_path.parent / f'results_{input_path.stem}.xlsx'
        
        log_path = None
        if args.verbose:
            log_path = input_path.parent / f'analysis_log_{input_path.stem}.txt'
        
        num_samples = process_data_and_calculate_titer(
            args.input,
            output_path,
            args.cutoff,
            args.method,
            args.replicates,
            args.verbose,
            log_path,
            args.extrapolation
        )
        
        print(f"Processing complete: {num_samples} samples analyzed")
        print(f"Results saved to {output_path}")
        if args.verbose:
            print(f"Analysis log saved to {log_path}")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}", file=sys.stderr)
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1
    
    return 0

if __name__ == '__main__':
    exit(main())
